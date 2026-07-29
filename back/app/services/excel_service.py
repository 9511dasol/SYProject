"""빈 Excel 템플릿에 DB 값을 채워 리포트를 만드는 서비스.

템플릿(app/assets/report_template.xlsx)은 **어느 달의 데이터도 갖고 있지 않다.**
수식·서식·레이아웃만 남기고 값은 전부 비운 상태이고, 시트 이름의 기간 자리에는
중립 토큰 PERIOD 가 들어 있다(`summary_PERIOD`, `네이버SA_PERIOD` …).
여기서는 그 토큰을 요청받은 기간으로 바꾸고 DB 값을 써 넣기만 한다.
템플릿을 다시 만들려면 `python -m scripts.reset_report_template` 을 쓴다.

예전에는 템플릿이 특정 달의 실적을 안고 있어서, 다른 달은 그 시트를 복사해 기간
문자열만 바꾸고 남은 숫자를 덮어쓰는 방식이었다. 덮어쓰지 못한 셀(매체별 전월 행,
네이버BS 브검MO 비용, 운영 메모 …)은 어느 달 파일에도 그대로 따라왔다.

시트 구조 (각 매체 시트 공통):
  col A   : 빈 스페이서
  col B   : 날짜
  row 7   : 컬럼 헤더
  row 8   : 전년동월  ┐ 같은 레이아웃. RAW 칸만 채우면
  row 9   : 전월      ┤ CTR·전환율 등 파생 지표는 템플릿 수식이 계산한다
  row 10  : 당월      ┘ (= 22행 TOTAL 참조)
  row 20  : ■ {매체명} 섹션 제목
  row 21  : 컬럼 헤더
  row 22  : TOTAL 집계 행
  row 23+ : 일별 RAW 데이터 (row 23 = 해당 월 1일)
"""

import calendar
import io
import math
import re
from datetime import date, datetime
from pathlib import Path

import openpyxl
import pandas as pd

TEMPLATE_PATH = Path(__file__).resolve().parent.parent / "assets" / "report_template.xlsx"

# 템플릿 시트 이름의 기간 자리표시자 — reset_report_template.py 와 같은 값이어야 한다.
PLACEHOLDER = "PERIOD"

SHEET_PREFIX: dict[str, str] = {
    "네이버SA":  "네이버SA",
    "네이버BS":  "네이버BS",
    "카카오SA":  "카카오SA",
    "구글SA":    "구글SA",
    "네이버PSA": "파워컨텐츠",  # DB 레이블 → 템플릿 시트 접두어 (기존 템플릿 호환)
}

SUMMARY_PREFIX = "summary"

DATA_ROW = 23   # row 23 = 해당 월 1일 (고정 오프셋)
YOY_ROW = 8     # 전년동월
PREV_ROW = 9    # 전월

_PERIOD_RE = re.compile(r"(\d{2,4})\s*년\s*(\d{1,2})\s*월")


# (kpi_field, col_number) — RAW 값 컬럼만 포함, 수식 컬럼 제외
# 네이버SA / 구글SA: 총전환수·전환율 등은 수식이므로 제외; 회원가입·구매완료·구매매출·신청만 입력
_NAVER_SA_COLS: list[tuple[str, int]] = [
    ("impressions",  3),   # 노출
    ("clicks",       4),   # 클릭
    ("cost",         7),   # 광고비(vat+)
    ("signup",      14),   # 회원가입
    ("purchase",    16),   # 구매완료
    ("revenue",     18),   # 구매매출
    ("apply",       21),   # 설명회신청
]

# 네이버BS: col7 광고비 = SUM(X:Y) 수식 → 원시 비용을 col24(브검PC)에 입력
# (col25 브검MO 는 DB에 구분이 없어 비워 둔다)
_NAVER_BS_COLS: list[tuple[str, int]] = [
    ("impressions",  3),
    ("clicks",       4),
    ("cost",        24),   # 브검PC (총비용 전부 — MO 구분 없음)
    ("signup",      14),
    ("purchase",    16),
    ("revenue",     18),
    ("apply",       21),
]

_KAKAO_SA_COLS: list[tuple[str, int]] = [
    ("impressions",  3),
    ("clicks",       4),
    ("cost",         7),
    ("signup",      11),   # 회원가입
    ("purchase",    13),   # 구매완료
    ("revenue",     15),   # 구매매출
]

# 네이버PSA(파워컨텐츠): 총전환수는 RAW 값 (수식 아님)
_POWER_COLS: list[tuple[str, int]] = [
    ("impressions",  3),
    ("clicks",       4),
    ("cost",         7),
    ("total_conv",   9),   # 총전환수 (RAW)
    ("signup",      12),
    ("purchase",    14),
    ("revenue",     16),
]

_GOOGLE_SA_COLS: list[tuple[str, int]] = [
    ("impressions",  3),
    ("clicks",       4),
    ("cost",         7),
    ("signup",      15),
    ("purchase",    17),
    ("revenue",     19),
    ("apply",       22),
]

_SHEET_COLS: dict[str, list[tuple[str, int]]] = {
    "네이버SA":  _NAVER_SA_COLS,
    "네이버BS":  _NAVER_BS_COLS,
    "카카오SA":  _KAKAO_SA_COLS,
    "네이버PSA": _POWER_COLS,
    "구글SA":    _GOOGLE_SA_COLS,
}


# 전년동월(8행)·전월(9행) 컬럼 — (field, col_number).
#
# 이 두 행은 당월(10행)과 레이아웃이 같고, 당월이 "=C22" 로 끌어오는 칸만 값이 필요하다.
# 나머지(CTR·전환율·markup 광고비 …)는 템플릿 수식이 같은 행 안에서 계산한다.
#
# field 의 "+" 는 합산을 뜻한다 — 총전환수를 당월 수식과 같은 정의로 맞추기 위함이다.
# (카카오 = 회원가입+구매완료, 구글 = 회원가입+구매완료+설명회신청, 파워 = DB 총전환수)
_AGGREGATE_COLS: dict[str, list[tuple[str, int]]] = {
    "네이버SA": [
        ("impressions", 3), ("clicks", 4), ("cost", 7),
        ("signup", 14), ("purchase", 16), ("revenue", 18), ("apply", 21),
    ],
    "네이버BS": [
        ("impressions", 3), ("clicks", 4), ("cost", 7),
        ("signup", 14), ("purchase", 16), ("revenue", 18), ("apply", 21),
    ],
    "카카오SA": [
        ("impressions", 3), ("clicks", 4), ("cost", 7), ("signup+purchase", 8),
        ("signup", 11), ("purchase", 13), ("revenue", 15),
    ],
    "구글SA": [
        ("impressions", 3), ("clicks", 4), ("cost", 7), ("signup+purchase+apply", 9),
        ("signup", 15), ("purchase", 17), ("revenue", 19), ("apply", 22),
    ],
    "네이버PSA": [
        ("impressions", 3), ("clicks", 4), ("cost", 7), ("conversions", 9),
        ("signup", 12), ("purchase", 14), ("revenue", 16),
    ],
}


# ── summary 시트 좌표 ─────────────────────────────────────────────────────
SUMMARY_END_DATE_CELL = "B1"    # 기간 종료일 (= 다음 달 1일). 잔여일수 계산의 기준
SUMMARY_BASE_DATE_CELL = "D1"   # 기준일. 원래 =TODAY() 였다 (아래 _base_date 주석 참고)
SUMMARY_MONTH_DAYS_CELL = "D3"
SUMMARY_COMMENT_CELL = "B32"

SUMMARY_YOY_DATE_ROW = 7    # 전년동월
PREV_MONTH_ROW = 8          # 전월
SUMMARY_CURR_DATE_ROW = 9   # 당월
SUMMARY_DAILY_ROW = 70      # 일별 구간 첫 행 (= 그 달 1일)
SUMMARY_DAILY_SLOTS = 31    # 70~100행. 짧은 달은 남는 칸을 비운다

_PREV_COL_MARKUP = 8  # 광고비(vat, markup) — DB에 markup 정보가 없어 비워 둔다

# ■ 매체별 예산 구간의 매체별 행 (D열 = 총합(vat,markup))
SUMMARY_BUDGET_COL = 4
SUMMARY_BUDGET_ROWS: dict[str, int] = {
    "네이버SA":  21,
    "카카오SA":  22,
    "구글SA":    23,
    "네이버BS":  25,
    "네이버PSA": 27,
}


def resolve_period(period: str, year: int | None, month: int | None) -> tuple[int, int]:
    """기간 라벨("26년 7월")에서 연·월을 얻는다. 명시된 값이 있으면 그쪽이 우선."""
    if year and month:
        return year, month
    m = _PERIOD_RE.search(period)
    if not m:
        raise ValueError(f"기간에서 연월을 읽을 수 없습니다: {period!r}")
    y, mo = int(m.group(1)), int(m.group(2))
    return (2000 + y if y < 100 else y), mo


def _next_month_first(year: int, month: int) -> date:
    return date(year + 1, 1, 1) if month == 12 else date(year, month + 1, 1)


def _base_date(year: int, month: int) -> date:
    """누적일수·잔여일수·Time elapsed·WoW 가 기준으로 삼는 날짜.

    템플릿은 원래 이 자리에 =TODAY() 를 썼다. 그래서 지난 달 리포트를 뽑으면 WoW 행이
    통째로 비고(그 달에 없는 날짜를 SUMIF 하므로) 예산소진율도 100%를 넘겨 표시됐다.
    끝난 달은 그 달의 마지막 날 다음 날을, 진행 중인 달은 오늘을 기준으로 삼는다 —
    같은 달을 언제 뽑아도 같은 파일이 나온다.
    """
    return min(date.today(), _next_month_first(year, month))


def _aggregate_value(totals: dict | None, field: str) -> float:
    """전년동월·전월 행에 넣을 값. 데이터가 없으면 0 — 빈칸으로 두면 YOY/MOM 이 조용히 빈다."""
    if not totals:
        return 0
    return sum(totals.get(part) or 0 for part in field.split("+"))


def _fill_aggregate_row(ws, media_label: str, row: int, totals: dict | None) -> None:
    for field, col in _AGGREGATE_COLS.get(media_label, []):
        ws.cell(row, col).value = _aggregate_value(totals, field)


def _fill_summary_dates(ws_sum, year: int, month: int) -> None:
    """summary 시트의 날짜 셀을 요청한 기간에 맞춰 다시 쓴다.

    ■ SA TOTAL 구간은 전년동월 / 전월 / 당월 3행이고, 아래 일별 구간은 그 달의 날짜다.
    전월(8행)은 _prev_month_cells 쪽에서 이미 채우므로 여기서는 나머지를 맡는다.
    """
    ws_sum.cell(SUMMARY_YOY_DATE_ROW, 2).value = datetime(year - 1, month, 1)
    ws_sum.cell(SUMMARY_CURR_DATE_ROW, 2).value = datetime(year, month, 1)

    days = calendar.monthrange(year, month)[1]
    for offset in range(SUMMARY_DAILY_SLOTS):
        # 30일 이하인 달은 남는 칸을 비운다 — 안 비우면 다음 달 1일이 그대로 남는다.
        # (매체 시트의 마지막 날짜 칸이 이 셀을 참조하므로 거기까지 함께 정리된다)
        ws_sum.cell(SUMMARY_DAILY_ROW + offset, 2).value = (
            datetime(year, month, offset + 1) if offset < days else None
        )


def _to_date(raw_date) -> date:
    """date / datetime / 'YYYY-MM-DD' 문자열을 Python date 객체로 변환"""
    if isinstance(raw_date, datetime):
        return raw_date.date()
    if isinstance(raw_date, date):
        return raw_date
    if isinstance(raw_date, str):
        return datetime.strptime(raw_date.strip()[:10], "%Y-%m-%d").date()
    if hasattr(raw_date, "date"):
        return raw_date.date()
    t = raw_date.timetuple()
    return date(t.tm_year, t.tm_mon, t.tm_mday)


def _clean(val) -> float | int | None:
    if val is None:
        return None
    if isinstance(val, float) and (math.isnan(val) or math.isinf(val)):
        return 0.0
    return val


_TEMPLATE_BYTES: bytes | None = None
_TEMPLATE_MTIME: float = 0.0


def _template_bytes() -> bytes:
    """템플릿 파일이 변경되면 자동으로 재읽기 (mtime 비교)"""
    global _TEMPLATE_BYTES, _TEMPLATE_MTIME
    mtime = TEMPLATE_PATH.stat().st_mtime
    if _TEMPLATE_BYTES is None or mtime != _TEMPLATE_MTIME:
        _TEMPLATE_BYTES = TEMPLATE_PATH.read_bytes()
        _TEMPLATE_MTIME = mtime
    return _TEMPLATE_BYTES


def _div(num: float, den: float) -> float:
    return num / den if den else 0.0


def _prev_month_cells(totals: dict) -> dict[int, float | None]:
    """DB 전월 합계 → summary '전월' 행의 {열 번호: 값}.

    이 행은 매체 시트와 달리 파생 지표까지 값으로 들어간다(템플릿에 수식이 없다).
    markup 광고비(8열)는 DB에 없어 None(빈칸)으로 둔다 — 모르는 값은 비우는 편이 낫다.
    """
    imp = totals["impressions"]
    clk = totals["clicks"]
    cost = totals["cost"]
    conv = totals["conversions"]
    rev = totals["revenue"]
    signup = totals["signup"]
    purchase = totals["purchase"]
    apply_ = totals["apply"]
    conv_ex = conv - apply_  # '신청제외' 전환 = 총전환 - 설명회신청

    return {
        3: imp,                      # 노출
        4: clk,                      # 클릭
        5: _div(clk, imp),           # CTR
        6: _div(cost, clk),          # CPC
        7: cost,                     # 광고비(vat+)
        _PREV_COL_MARKUP: None,      # 광고비(vat, markup)
        9: conv,                     # 총전환수
        10: _div(conv, clk),         # 전환율
        11: _div(cost, conv),        # 전환단가
        12: conv_ex,                 # 총전환수(신청제외)
        13: _div(conv_ex, clk),      # 전환율(신청제외)
        14: _div(cost, conv_ex),     # 전환단가(신청제외)
        15: signup,                  # 회원가입
        16: _div(signup, clk),       # 회원가입률
        17: purchase,                # 구매완료
        18: _div(purchase, clk),     # 결제완료율
        19: rev,                     # 구매매출
        20: _div(rev, cost),         # 구매수익률(ROAS)
        21: _div(rev, purchase),     # 구매단가
        22: apply_,                  # 설명회신청
    }


def _retitle(wb, period: str) -> None:
    """시트 이름과 수식 안의 PERIOD 토큰을 요청한 기간으로 바꾼다.

    수식은 '{접두어}_PERIOD'!C23 형태로 시트를 참조하므로, 이름만 바꾸면 참조가 깨진다.
    두 작업은 항상 같이 해야 한다.
    """
    token = f"_{PLACEHOLDER}"
    new = f"_{period}"

    for name in list(wb.sheetnames):
        if name.endswith(token):
            wb[name].title = f"{name[: -len(token)]}{new}"

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("=") and token in cell.value:
                    cell.value = cell.value.replace(token, new)


class ExcelService:
    def fill_template(
        self,
        media_kpis: dict[str, pd.DataFrame],
        period: str,
        year: int | None = None,
        month: int | None = None,
        prev_totals: dict | None = None,
        yoy_media_totals: dict[str, dict] | None = None,
        *,
        prev_media_totals: dict[str, dict] | None = None,
        media_budgets: dict[str, float] | None = None,
        comment: str | None = None,
    ) -> bytes:
        """빈 템플릿에 DB 값을 채워 넣고 xlsx bytes 를 반환한다.

        media_kpis        : {매체: 일별 KPI DataFrame} — 매체 시트 23행 이하
        period            : 시트 이름에 쓸 기간 라벨 ("26년 7월")
        year, month       : 생략하면 period 에서 읽는다
        prev_totals       : 직전월 전체 합계 → summary '전월' 행
        yoy_media_totals  : 1년 전 같은 달의 매체별 합계 → 매체 시트 8행
        prev_media_totals : 직전월 매체별 합계 → 매체 시트 9행
        media_budgets     : {매체: 월 예산} → summary ■ 매체별 예산
        comment           : summary B32 코멘트

        데이터가 없는 매체·기간은 0으로 채운다. 빈칸으로 두면 그 값을 쓰는 YOY/MOM 행이
        조용히 비어버려서, 받는 사람은 '숫자가 없다'와 '0이다'를 구분할 수 없다.
        """
        year, month = resolve_period(period, year, month)

        wb = openpyxl.load_workbook(
            io.BytesIO(_template_bytes()),
            keep_links=True,   # False로 하면 수식 참조가 끊겨 Excel이 손상 경고를 표시함
            keep_vba=False,
        )
        _retitle(wb, period)

        summary_name = f"{SUMMARY_PREFIX}_{period}"
        if summary_name not in wb.sheetnames:
            raise ValueError(
                f"템플릿에 {SUMMARY_PREFIX}_{PLACEHOLDER} 시트가 없습니다. "
                "python -m scripts.reset_report_template 으로 템플릿을 다시 만들어 주세요."
            )

        self._fill_summary(
            wb[summary_name], year, month, prev_totals, media_budgets, comment
        )

        filled = 0
        for media_label, prefix in SHEET_PREFIX.items():
            sheet_name = f"{prefix}_{period}"
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]

            df = media_kpis.get(media_label)
            if df is not None:
                self._fill_sheet(ws, df, _SHEET_COLS.get(media_label, _NAVER_SA_COLS))
                filled += 1

            # 전년동월·전월은 데이터가 없는 매체도 0으로 채운다
            _fill_aggregate_row(ws, media_label, YOY_ROW, (yoy_media_totals or {}).get(media_label))
            _fill_aggregate_row(ws, media_label, PREV_ROW, (prev_media_totals or {}).get(media_label))

        if filled == 0:
            raise ValueError(f"'{period}' 기간에 채울 매체 데이터가 없습니다.")

        wb.calculation.forceFullCalc = True  # 열 때 Excel이 수식 전체 재계산
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    def _fill_summary(
        self,
        ws_sum,
        year: int,
        month: int,
        prev_totals: dict | None,
        media_budgets: dict[str, float] | None,
        comment: str | None,
    ) -> None:
        end = _next_month_first(year, month)
        base = _base_date(year, month)
        ws_sum[SUMMARY_END_DATE_CELL] = datetime(end.year, end.month, end.day)
        ws_sum[SUMMARY_BASE_DATE_CELL] = datetime(base.year, base.month, base.day)
        ws_sum[SUMMARY_MONTH_DAYS_CELL] = calendar.monthrange(year, month)[1]

        # 전년동월·당월·일별 날짜 — 요청한 기간의 날짜로 채운다
        _fill_summary_dates(ws_sum, year, month)

        # 값 대입은 .value 로 한다 — ws.cell(r, c, None) 은 openpyxl이 무시해서
        # 빈칸 처리가 안 된다.
        prev_year, prev_month = (year - 1, 12) if month == 1 else (year, month - 1)
        ws_sum.cell(PREV_MONTH_ROW, 2).value = datetime(prev_year, prev_month, 1)
        cells = _prev_month_cells(prev_totals) if prev_totals else {}
        for col in range(3, 23):
            ws_sum.cell(PREV_MONTH_ROW, col).value = cells.get(col)

        # ■ 매체별 예산 — 설정하지 않은 매체는 0 (예산소진율은 IFERROR 로 비어 보인다)
        budgets = media_budgets or {}
        for media_label, row in SUMMARY_BUDGET_ROWS.items():
            ws_sum.cell(row, SUMMARY_BUDGET_COL).value = budgets.get(media_label) or 0

        ws_sum[SUMMARY_COMMENT_CELL] = comment or None

    def _fill_sheet(
        self,
        ws,
        df: pd.DataFrame,
        col_map: list[tuple[str, int]],
    ) -> None:
        """row 23 = 1일 오프셋 기준으로 날짜별 RAW 지표 기입.
        B열(날짜)은 수식 참조 대신 실제 날짜 값을 직접 기입한다.
        """
        fields = [f for f, _ in col_map]
        cols = [c for _, c in col_map]
        for field in fields:
            if field not in df.columns:
                df[field] = 0.0
        dates = df["date"].tolist()
        values = df[fields].values
        for i, date_val in enumerate(dates):
            d = _to_date(date_val)
            target_row = DATA_ROW + d.day - 1
            ws.cell(target_row, 2, d)          # B열: 날짜 직접 기입
            row_vals = values[i]
            for col_num, val in zip(cols, row_vals, strict=True):
                ws.cell(target_row, col_num, _clean(val))
