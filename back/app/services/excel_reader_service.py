"""기존 Excel 파일을 읽어 JSON으로 반환하는 서비스."""

import io
import math
from datetime import datetime

import openpyxl


def _safe(v, default: float = 0.0) -> float:
    if v is None:
        return default
    try:
        f = float(v)
        return default if (math.isnan(f) or math.isinf(f)) else f
    except (TypeError, ValueError):
        return default


def _date_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    return str(v)[:10]


def _cell_val(ws, row: int, col: int):
    v = ws.cell(row, col).value
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
        return 0.0
    return v


# summary 시트 row 6 컬럼 순서 (col 2~21, 0-indexed offset 0~19)
SA_TOTAL_COLS = [
    "date", "impressions", "clicks", "ctr", "cpc",
    "cost_vat", "cost_markup", "total_conv", "conv_rate", "conv_cost",
    "total_conv_ex", "conv_rate_ex", "conv_cost_ex",
    "signup", "signup_rate", "purchase", "purchase_rate",
    "revenue", "roas", "revenue_per_purchase",
]


class ExcelReaderService:
    # 시트명 → DB 레이블 (기존 "파워컨텐츠" 시트를 "네이버PSA"로 읽음)
    SHEET_TO_LABEL: dict[str, str] = {
        "네이버SA":  "네이버SA",
        "네이버BS":  "네이버BS",
        "카카오SA":  "카카오SA",
        "구글SA":    "구글SA",
        "네이버PSA": "네이버PSA",
        "파워컨텐츠": "네이버PSA",  # 기존 템플릿 시트명 → 새 DB 레이블
    }

    def read_report(self, excel_bytes: bytes) -> dict:
        # read_only=True: 스트리밍 파싱으로 메모리 절약 (시트 100개 이상 대용량 파일 대응)
        # data_only=True: 수식 셀을 마지막 계산된 캐시 값으로 읽음
        #   → 날짜 수식(='summary_...'!B70)도 datetime으로 정상 반환되어 row[0] 필터를 통과함
        wb = openpyxl.load_workbook(
            io.BytesIO(excel_bytes), read_only=True, data_only=True
        )
        try:
            period = self._detect_period(wb)
            sname = f"summary_{period}"
            ws = wb[sname] if sname in wb.sheetnames else None
            if ws is None:
                raise ValueError(f"'{sname}' 시트를 찾을 수 없습니다.")

            media = {}
            for sheet_prefix, db_label in self.SHEET_TO_LABEL.items():
                ms = f"{sheet_prefix}_{period}"
                if ms in wb.sheetnames and db_label not in media:
                    media[db_label] = self._parse_media_sheet(wb[ms])

            return {
                "period": period,
                "period_info": self._parse_period_info(ws),
                "sa_total": self._parse_sa_total(ws),
                "budget_table": self._parse_budget_table(ws),
                "comment": self._parse_comment(ws),
                "daily_total": self._parse_daily_total(ws),
                "media": media,
            }
        finally:
            wb.close()

    # ── 내부 헬퍼 ─────────────────────────────────────────────────────────────

    def _detect_period(self, wb) -> str:
        for prefix in ("summary_26년", "summary_25년", "summary_"):
            candidates = [s for s in wb.sheetnames if s.startswith(prefix)]
            if candidates:
                return candidates[-1].replace("summary_", "")
        raise ValueError("summary 시트를 찾을 수 없습니다.")

    def _parse_period_info(self, ws) -> dict:
        return {
            "remaining_days": int(_safe(ws.cell(3, 2).value)),
            "elapsed_days": int(_safe(ws.cell(3, 3).value)),
            "total_days": int(_safe(ws.cell(3, 4).value)),
        }

    def _parse_sa_total(self, ws) -> dict:
        """summary 시트 rows 7~14: 전년/전월/당월/YOY/MOM/전주/금주/WoW"""
        # 컬럼 헤더는 row 6 (col 2~21)
        headers = [
            str(ws.cell(6, c).value or "").replace("\n", " ")
            for c in range(2, 22)
        ]

        def _parse_comparison_row(row_num: int, label: str) -> dict:
            row: dict = {"label": label}
            for i, key in enumerate(SA_TOTAL_COLS):
                v = ws.cell(row_num, 2 + i).value
                if key == "date":
                    row[key] = _date_str(v) if isinstance(v, datetime) else str(v or "")
                else:
                    row[key] = _safe(v)
            return row

        return {
            "headers": headers,
            "rows": [
                _parse_comparison_row(7, "전년"),
                _parse_comparison_row(8, "전월"),
                _parse_comparison_row(9, "당월"),
                _parse_comparison_row(10, "YOY"),
                _parse_comparison_row(11, "MOM"),
                _parse_comparison_row(12, "전주"),
                _parse_comparison_row(13, "금주"),
                _parse_comparison_row(14, "WoW"),
            ],
        }

    def _parse_budget_table(self, ws) -> list[dict]:
        rows = []
        for r in range(21, 29):
            category = ws.cell(r, 3).value
            if not category:
                continue
            rows.append({
                "category": str(category),
                "budget": _safe(ws.cell(r, 4).value),
                "spent": _safe(ws.cell(r, 5).value),
                "burn_rate": _safe(ws.cell(r, 6).value),
                "impressions": _safe(ws.cell(r, 10).value),
                "clicks": _safe(ws.cell(r, 11).value),
                "cost_vat": _safe(ws.cell(r, 12).value),
                "total_conv": _safe(ws.cell(r, 13).value),
                "conv_rate": _safe(ws.cell(r, 14).value),
                "conv_cost": _safe(ws.cell(r, 15).value),
            })
        return rows

    def _parse_comment(self, ws) -> str:
        v = ws.cell(32, 2).value
        return str(v) if v else ""

    def _parse_daily_total(self, ws) -> list[dict]:
        rows = []
        for r in range(70, 102):
            date_val = ws.cell(r, 2).value
            if date_val is None:
                continue
            imp = _safe(ws.cell(r, 3).value)
            clk = _safe(ws.cell(r, 4).value)
            if imp == 0 and clk == 0:
                continue
            rows.append({
                "date": _date_str(date_val),
                "impressions": imp,
                "clicks": clk,
                "ctr": _safe(ws.cell(r, 5).value),
                "cpc": _safe(ws.cell(r, 6).value),
                "cost": _safe(ws.cell(r, 7).value),
                "total_conv": _safe(ws.cell(r, 9).value),
                "conv_rate": _safe(ws.cell(r, 10).value),
                "conv_cost": _safe(ws.cell(r, 11).value),
            })
        return rows

    def _parse_media_sheet(self, ws) -> dict:
        # Read all candidate headers (cols 2–34), replacing None with "".
        # Do NOT stop at the first None — some templates have sparse headers.
        raw: list[str] = []
        for c in range(2, 35):
            v = ws.cell(21, c).value
            raw.append("" if v is None else str(v).replace("\n", " "))
        # Drop trailing empty entries so len(headers) == useful column count.
        while raw and raw[-1] == "":
            raw.pop()
        headers = raw

        n = len(headers)
        total = self._parse_row(ws, 22, n)
        daily = []
        for r in range(23, 54):
            row = self._parse_row(ws, r, n)
            if not row[0] or (_safe(row[1]) == 0 and _safe(row[2]) == 0):
                continue
            daily.append(row)

        return {"headers": headers, "total": total, "daily": daily}

    def _parse_row(self, ws, row_num: int, ncols: int) -> list:
        return [_cell_val(ws, row_num, 2 + i) for i in range(ncols)]

    def to_db_dataframe(self, report: dict):
        """media 시트 데이터를 DB 저장용 DataFrame으로 변환"""
        import pandas as pd
        records = []
        for media_label, sheet in report["media"].items():
            headers = sheet["headers"]

            def _find(predicate, default):
                return next((i for i, h in enumerate(headers) if predicate(h)), default)

            col_imp = 1
            col_clk = 2
            col_cost = _find(lambda h: "광고비" in h and "vat" in h.lower(), 5)
            col_conv = _find(lambda h: h.startswith("총전환수") and "제외" not in h, 6)
            col_rev = _find(lambda h: "구매매출" in h, 16)
            col_signup = _find(lambda h: "회원가입" in h, None)
            col_purchase = _find(lambda h: "구매완료" in h and "매출" not in h, None)
            col_apply = _find(lambda h: "신청" in h and "회원" not in h, None)

            def _get(row, col):
                return _safe(row[col] if col is not None and col < len(row) else None)

            for row in sheet["daily"]:
                date_val = row[0]
                if not date_val:
                    continue
                records.append({
                    "report_date": date_val,
                    "campaign_type": media_label,
                    "impressions": int(_get(row, col_imp)),
                    "clicks": int(_get(row, col_clk)),
                    "cost": _get(row, col_cost),
                    "conversions": int(_get(row, col_conv)),
                    "conversion_revenue": _get(row, col_rev),
                    "signup": _get(row, col_signup),
                    "purchase": _get(row, col_purchase),
                    "apply": _get(row, col_apply),
                })
        return pd.DataFrame(records)
