"""report_template.xlsx 를 '데이터 없는 빈 템플릿'으로 되돌린다.

지금까지 템플릿은 특정 달(26년 7월)의 실적이 그대로 박힌 파일이었다. 그래서 export 는
"템플릿에 있는 달의 시트를 복사해서 기간 문자열만 바꾸고, 남은 숫자를 DB 값으로 덮어쓴다"
라는 방식이었는데, 덮어쓰지 못한 셀(매체별 전월 행, 브검MO 비용, 운영 메모 …)은 어느 달
파일을 뽑아도 7월 숫자가 그대로 따라왔다.

이 스크립트는 그 파일에서 **수식·서식·레이아웃은 전부 남기고 값만 비운** 템플릿을 만든다.
결과물은 어느 달에도 속하지 않으므로 시트 이름의 기간도 중립 토큰(PERIOD)으로 바꾼다.
ExcelService 는 이 토큰을 요청받은 기간으로 치환해서 쓴다 — 더 이상 남의 달 시트를
복사하지 않는다.

    python -m scripts.reset_report_template            # 템플릿을 제자리에서 갱신
    python -m scripts.reset_report_template out.xlsx   # 결과를 다른 경로로

되돌리려면 git 으로 report_template.xlsx 를 복원하면 된다.
"""

import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.cell.cell import MergedCell

_ASSETS = Path(__file__).resolve().parent.parent / "app" / "assets"
TEMPLATE_PATH = _ASSETS / "report_template.xlsx"

# 시트 이름에 남길 기간 자리표시자. ASCII 라서 한글 텍스트 셀과 우연히 겹칠 일이 없다.
PLACEHOLDER = "PERIOD"

SUMMARY_PREFIX = "summary"

# ── 매체 시트 좌표 ─────────────────────────────────────────────────────────
MEDIA_BALANCE_ROW = 5      # 광고잔액 등 상단 요약 (DB에 없는 값 → 0)
MEDIA_YOY_ROW = 8          # 전년동월
MEDIA_PREV_ROW = 9         # 전월
MEDIA_CURR_ROW = 10        # 당월 — 8·9행의 수식 골격을 여기서 가져온다
MEDIA_LASTWEEK_ROW = 17    # 전주 (일부 시트에 값으로 굳은 셀이 있다)
MEDIA_THISWEEK_ROW = 18
MEDIA_DATA_FIRST = 23      # 일별 RAW 첫 행 (= 1일)
MEDIA_DATA_LAST = 53       # 31일 + 여유

# ── summary 시트 좌표 ─────────────────────────────────────────────────────
SUMMARY_END_DATE = "B1"     # 기간 종료일(= 다음 달 1일)
SUMMARY_BASE_DATE = "D1"    # 기준일 — 원래 =TODAY() 였다
SUMMARY_MONTH_DAYS = "D3"
SUMMARY_YOY_ROW = 7
SUMMARY_PREV_ROW = 8
SUMMARY_CURR_ROW = 9
SUMMARY_BUDGET_ROWS = range(21, 28)   # ■ 매체별 예산
SUMMARY_BUDGET_COL = 4                # D열 총합(vat,markup)
SUMMARY_COMMENT = "B32"
SUMMARY_DAILY_FIRST = 70
SUMMARY_DAILY_LAST = 100

# "=C22" 처럼 22행(TOTAL) 한 칸만 가리키는 수식 = DB 값이 들어갈 자리
_RAW_SLOT_RE = re.compile(r"^=\$?[A-Z]{1,3}\$?22$")
_CELL_REF_RE = re.compile(r"(\$?[A-Z]{1,3}\$?)(\d+)")
_ANY_REF_RE = re.compile(r"\$?[A-Z]{1,3}\$?\d+")


def _is_formula(value) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _put(ws, row: int, col: int, value) -> None:
    """병합 셀의 종속 칸은 값을 가질 수 없다 — 건너뛴다."""
    cell = ws.cell(row, col)
    if not isinstance(cell, MergedCell):
        cell.value = value


def _shift_row(formula: str, src_row: int, dst_row: int) -> str | None:
    """수식의 셀 참조 행 번호를 src_row → dst_row 로 옮긴다.

    src_row 이외의 행을 가리키는 참조가 하나라도 있으면 None 을 돌려준다 —
    그런 수식은 '같은 행 안에서 계산되는 파생 지표'가 아니라서 그대로 옮기면 틀린다.
    (예: =SUM(C36:C38) 은 데이터 구간을 가리키므로 대상이 아니다)
    """
    refs = _CELL_REF_RE.findall(formula)
    if not refs or any(int(row) != src_row for _, row in refs):
        return None
    return _CELL_REF_RE.sub(lambda m: f"{m.group(1)}{dst_row}", formula)


def _populated_cols(ws, row: int) -> list[int]:
    return [c for c in range(3, ws.max_column + 1) if ws.cell(row, c).value is not None]


def _retitle_sheets(wb, token: str) -> None:
    """시트 이름과 수식 안의 기간 문자열을 중립 토큰으로 바꾼다."""
    for name in list(wb.sheetnames):
        if name.endswith(f"_{token}"):
            wb[name].title = f"{name[: -len(token)]}{PLACEHOLDER}"

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if _is_formula(cell.value) and token in cell.value:
                    cell.value = cell.value.replace(token, PLACEHOLDER)


def _rewire_base_date(wb) -> None:
    """TODAY() 를 summary 의 '기준일' 셀 참조로 바꾼다.

    누적일수·잔여일수·Time elapsed·WoW(전주 대비)가 전부 TODAY() 기준이라, 지난 달
    리포트를 뽑으면 WoW 행이 통째로 비고 예산소진율도 100%를 넘겨 표시됐다. 기준일을
    한 셀로 모아 두면 ExcelService 가 '그 달의 마지막 날(진행 중이면 오늘)'을 넣어 줄 수
    있고, 같은 달은 언제 뽑아도 같은 파일이 나온다.
    """
    summary = wb[f"{SUMMARY_PREFIX}_{PLACEHOLDER}"]
    summary[SUMMARY_BASE_DATE] = None  # 자기 자신을 참조하지 않도록 먼저 비운다

    for ws in wb.worksheets:
        ref = (
            f"${SUMMARY_BASE_DATE[0]}${SUMMARY_BASE_DATE[1:]}"
            if ws is summary
            else f"'{SUMMARY_PREFIX}_{PLACEHOLDER}'!${SUMMARY_BASE_DATE[0]}${SUMMARY_BASE_DATE[1:]}"
        )
        for row in ws.iter_rows():
            for cell in row:
                if _is_formula(cell.value) and "TODAY()" in cell.value:
                    cell.value = cell.value.replace("TODAY()", ref)


def _reset_media_sheet(ws) -> None:
    # 광고잔액 등 상단 요약 — DB에 없는 값이라 0으로 둔다 (매체명 같은 라벨은 남긴다)
    for col in range(3, ws.max_column + 1):
        if isinstance(ws.cell(MEDIA_BALANCE_ROW, col).value, (int, float)):
            _put(ws, MEDIA_BALANCE_ROW, col, 0)

    # 전년동월·전월 행을 '당월' 행과 같은 골격으로 다시 세운다.
    #
    # 원래 8행은 외부 통합문서(=[1]네이버SA_7월!C22)를 참조했고 9행은 통째로 상수였다.
    # 받는 사람에게 그 외부 파일이 없어 8행은 빈칸이 되고, 9행은 어느 달을 뽑아도
    # 템플릿을 만든 달의 숫자가 그대로 남아 매체별 MOM 이 조용히 틀렸다.
    # 당월(10행)은 같은 레이아웃을 수식으로 온전히 갖고 있으므로 그것을 본으로 삼는다.
    shaped = set(_populated_cols(ws, MEDIA_CURR_ROW))
    for col in shaped:
        current = ws.cell(MEDIA_CURR_ROW, col).value
        for target_row in (MEDIA_YOY_ROW, MEDIA_PREV_ROW):
            if _is_formula(current) and not _RAW_SLOT_RE.match(current):
                # 파생 지표(CTR·전환율·markup 광고비 …) — 같은 행을 가리키게 옮겨 심는다
                _put(ws, target_row, col, _shift_row(current, MEDIA_CURR_ROW, target_row))
            else:
                # =C22 자리 = DB 합계가 들어갈 칸
                _put(ws, target_row, col, 0)

    # 당월 행에 짝이 없는 칸에 남은 옛 숫자 (템플릿을 만들 때 손으로 넣은 값)
    for target_row in (MEDIA_YOY_ROW, MEDIA_PREV_ROW):
        for col in range(3, ws.max_column + 1):
            if col not in shaped and not _is_formula(ws.cell(target_row, col).value):
                _put(ws, target_row, col, None)

    # 전주(17행)에 값으로 굳어 있는 셀 — 금주(18행) 수식을 한 행 위로 옮겨 되살린다
    for col in range(3, ws.max_column + 1):
        if _is_formula(ws.cell(MEDIA_LASTWEEK_ROW, col).value):
            continue
        sibling = ws.cell(MEDIA_THISWEEK_ROW, col).value
        if not _is_formula(sibling):
            continue
        restored = _shift_row(sibling, MEDIA_THISWEEK_ROW, MEDIA_LASTWEEK_ROW)
        if restored:
            _put(ws, MEDIA_LASTWEEK_ROW, col, restored)

    # 일별 RAW 구간 — 값을 비운다.
    #
    # 수식이라도 셀 참조가 없으면(카카오 광고비의 "=70*1.1" 처럼) 모양만 수식인 손입력
    # 값이다. 남겨두면 데이터가 없는 날에 유령 실적으로 살아남으므로 같이 지운다.
    for row in range(MEDIA_DATA_FIRST, MEDIA_DATA_LAST + 1):
        for col in range(3, ws.max_column + 1):
            value = ws.cell(row, col).value
            if value is None:
                continue
            if not _is_formula(value) or not _ANY_REF_RE.search(value):
                _put(ws, row, col, None)


def _reset_summary_sheet(ws) -> None:
    ws[SUMMARY_END_DATE] = None
    ws[SUMMARY_MONTH_DAYS] = None
    ws[SUMMARY_COMMENT] = None

    # ■ SA TOTAL 의 전년동월·전월·당월 날짜와, 상수로 박힌 전월 실적
    for row in (SUMMARY_YOY_ROW, SUMMARY_PREV_ROW, SUMMARY_CURR_ROW):
        _put(ws, row, 2, None)
    for col in range(3, ws.max_column + 1):
        if not _is_formula(ws.cell(SUMMARY_PREV_ROW, col).value):
            _put(ws, SUMMARY_PREV_ROW, col, None)

    # ■ 매체별 예산 — DB(marketing_period_meta.media_budgets)에서 채운다
    for row in SUMMARY_BUDGET_ROWS:
        if not _is_formula(ws.cell(row, SUMMARY_BUDGET_COL).value):
            _put(ws, row, SUMMARY_BUDGET_COL, 0)

    # 일별 구간 날짜
    for row in range(SUMMARY_DAILY_FIRST, SUMMARY_DAILY_LAST + 1):
        _put(ws, row, 2, None)


def reset_template(src: Path, dst: Path) -> None:
    wb = openpyxl.load_workbook(src, keep_links=True, keep_vba=False)

    token = next(
        (n[len(SUMMARY_PREFIX) + 1 :] for n in wb.sheetnames if n.startswith(f"{SUMMARY_PREFIX}_")),
        None,
    )
    if token is None:
        raise SystemExit(f"summary_* 시트를 찾을 수 없습니다: {wb.sheetnames}")
    if token != PLACEHOLDER:
        _retitle_sheets(wb, token)

    _rewire_base_date(wb)

    summary_name = f"{SUMMARY_PREFIX}_{PLACEHOLDER}"
    for ws in wb.worksheets:
        if ws.title == summary_name:
            _reset_summary_sheet(ws)
        else:
            _reset_media_sheet(ws)

    # 외부 통합문서 참조는 위에서 전부 걷어냈다 — 링크 정의만 남으면 Excel이 열 때마다
    # "연결된 데이터를 업데이트할까요?"를 묻는다.
    wb._external_links = []
    wb.active = 0

    wb.save(dst)
    print(f"기간 토큰 {token!r} → {PLACEHOLDER!r}, 시트 {len(wb.sheetnames)}개 초기화 → {dst}")


def main(argv: list[str]) -> None:
    # 제자리 갱신이 기본이다 — 원본은 git 이 갖고 있으므로 따로 백업하지 않는다.
    reset_template(TEMPLATE_PATH, Path(argv[1]).resolve() if len(argv) > 1 else TEMPLATE_PATH)


if __name__ == "__main__":
    main(sys.argv)
