"""report_template.xlsx에 새 기간 시트를 추가합니다.

기존 기간의 모든 시트(매체 + summary)를 복사하고,
수식 안의 기간 문자열도 새 기간으로 교체합니다.

사용법:
    python back/scripts/add_template_period.py "26년 5월" "26년 6월"
"""

import sys
from pathlib import Path

import openpyxl

TEMPLATE_PATH = (
    Path(__file__).resolve().parent.parent / "example" / "report_template.xlsx"
)


def add_period_sheets(source_period: str, target_period: str) -> None:
    wb = openpyxl.load_workbook(TEMPLATE_PATH, keep_links=True, keep_vba=False)

    source_sheets = [s for s in wb.sheetnames if source_period in s]
    if not source_sheets:
        print(f"[오류] '{source_period}' 포함 시트를 찾을 수 없습니다.")
        print(f"현재 시트 목록: {wb.sheetnames}")
        sys.exit(1)

    added = 0
    for sheet_name in source_sheets:
        new_name = sheet_name.replace(source_period, target_period)
        if new_name in wb.sheetnames:
            print(f"이미 존재: {new_name} (건너뜀)")
            continue

        new_ws = wb.copy_worksheet(wb[sheet_name])
        new_ws.title = new_name

        # 수식 안의 source_period 참조를 target_period로 교체
        for row in new_ws.iter_rows():
            for cell in row:
                if (
                    cell.value
                    and isinstance(cell.value, str)
                    and cell.value.startswith("=")
                    and source_period in cell.value
                ):
                    cell.value = cell.value.replace(source_period, target_period)

        print(f"추가됨: {sheet_name}  →  {new_name}")
        added += 1

    if added == 0:
        print("새로 추가된 시트가 없습니다.")
        return

    wb.save(TEMPLATE_PATH)
    print(f"\n{added}개 시트 추가 완료 → {TEMPLATE_PATH}")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("사용법: python add_template_period.py <소스기간> <대상기간>")
        print("예시:   python add_template_period.py '26년 5월' '26년 6월'")
        sys.exit(1)

    add_period_sheets(sys.argv[1], sys.argv[2])
