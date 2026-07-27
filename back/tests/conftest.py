"""테스트 공통 픽스처.

외부 의존(DB/LLM/메일/스토리지) 없이 순수 로직·엑셀 파싱만 검증하는 1단계 테스트용.
엑셀 픽스처는 바이너리 파일을 커밋하지 않고 openpyxl/pandas로 즉석 생성한다
(포맷이 바뀌면 여기만 고치면 되고, git diff 도 읽을 수 있다)."""

import io

import openpyxl
import pandas as pd
import pytest


def _df_to_xlsx_bytes(sheets: dict[str, pd.DataFrame]) -> bytes:
    """{시트명: DataFrame(header 없이 raw 셀 배치)} → xlsx 바이트."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for name, df in sheets.items():
            df.to_excel(writer, sheet_name=name, header=False, index=False)
    return buf.getvalue()


@pytest.fixture
def make_keyword_xlsx():
    """keyword_compare 용 xlsx 생성기. rows 는 header 없는 2D 리스트."""
    def _make(sheets: dict[str, list[list]]) -> bytes:
        return _df_to_xlsx_bytes({n: pd.DataFrame(rows) for n, rows in sheets.items()})
    return _make


@pytest.fixture
def make_report_xlsx():
    """excel_reader.read_report 용 최소 summary 워크북 생성기.

    셀 좌표를 직접 채워 실제 리포트 파일 구조를 최소로 재현한다.
    cells: {시트명: {(row, col): value}}
    """
    def _make(period: str, cells: dict[str, dict[tuple[int, int], object]]) -> bytes:
        wb = openpyxl.Workbook()
        # 기본 생성 시트 제거
        wb.remove(wb.active)
        for sheet_name, cell_map in cells.items():
            ws = wb.create_sheet(title=sheet_name)
            for (row, col), value in cell_map.items():
                ws.cell(row=row, column=col, value=value)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    return _make
