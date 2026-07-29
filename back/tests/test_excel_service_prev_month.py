"""빈 템플릿 + DB 값으로 리포트 엑셀을 만드는 경로.

템플릿(app/assets/report_template.xlsx)은 어느 달의 데이터도 갖고 있지 않다.
수식·서식만 남기고 값은 비었고, 시트 이름의 기간 자리에는 PERIOD 토큰이 들어 있다.
그래서 여기서 검증할 것은 "채워 넣은 값이 맞는가"와 "채우지 않은 칸에 남의 달 숫자가
남지 않는가" 두 가지다.

예전에는 템플릿이 특정 달의 실적을 안고 있고 다른 달은 그 시트를 복사해 만들었다.
덮어쓰지 못한 셀(매체별 전월 행, 네이버BS 브검MO 비용, 운영 메모 …)이 어느 달 파일에도
그대로 따라왔던 것이 이 구조를 바꾼 이유다.
"""

import io
import re
import zipfile
from datetime import datetime

import openpyxl
import pandas as pd
import pytest

from app.services.excel_service import (
    PLACEHOLDER,
    PREV_MONTH_ROW,
    PREV_ROW,
    SUMMARY_BUDGET_COL,
    SUMMARY_BUDGET_ROWS,
    SUMMARY_DAILY_ROW,
    SUMMARY_DAILY_SLOTS,
    TEMPLATE_PATH,
    YOY_ROW,
    ExcelService,
    _prev_month_cells,
    resolve_period,
)

# 실제 6월 실적 — 파생 지표 검산에 쓴다
PREV_TOTALS = {
    "impressions": 4574572,
    "clicks": 103847,
    "cost": 36887933.0,
    "conversions": 7915,
    "revenue": 1015564704.0,
    "signup": 1497.0,
    "purchase": 5298.0,
    "apply": 1119.0,
}

_ALL_MEDIA = ("네이버SA", "네이버BS", "카카오SA", "구글SA", "네이버PSA")

_SAMPLE = {
    "impressions": 111, "clicks": 22, "cost": 3300.0, "conversions": 44,
    "revenue": 5.0, "signup": 6.0, "purchase": 7.0, "apply": 8.0,
}


def _kpi_frame(year: int, month: int, days: int) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "date": f"{year}-{month:02d}-{day:02d}",
                "impressions": 100 * day,
                "clicks": 10 * day,
                "cost": 1000 * day,
                "total_conv": day,
                "signup": day,
                "purchase": day,
                "revenue": 5000 * day,
                "apply": day,
            }
            for day in range(1, days + 1)
        ]
    )


def _export(period: str, year: int, month: int, days: int, **kwargs):
    kpis = {label: _kpi_frame(year, month, days) for label in _ALL_MEDIA}
    raw = ExcelService().fill_template(kpis, period, year, month, **kwargs)
    return openpyxl.load_workbook(io.BytesIO(raw))


needs_template = pytest.mark.skipif(not TEMPLATE_PATH.exists(), reason="report_template.xlsx 없음")


# ── 기간 라벨 파싱 ────────────────────────────────────────────────────────────


class TestResolvePeriod:
    def test_reads_year_and_month_from_label(self):
        assert resolve_period("26년 7월", None, None) == (2026, 7)

    def test_four_digit_year_is_kept(self):
        assert resolve_period("2026년 12월", None, None) == (2026, 12)

    def test_explicit_arguments_win(self):
        assert resolve_period("26년 7월", 2025, 3) == (2025, 3)

    def test_unparsable_label_raises(self):
        with pytest.raises(ValueError):
            resolve_period("지난달", None, None)


# ── summary 전월 행 (여기만 파생 지표까지 값으로 들어간다) ────────────────────


class TestPrevMonthCells:
    def test_derived_metrics_match_template_figures(self):
        cells = _prev_month_cells(PREV_TOTALS)

        assert cells[3] == 4574572                      # 노출
        assert cells[4] == 103847                       # 클릭
        assert round(cells[5], 4) == 0.0227             # CTR
        assert round(cells[6]) == 355                   # CPC
        assert cells[7] == 36887933.0                   # 광고비(vat+)
        assert round(cells[10], 4) == 0.0762            # 전환율
        assert round(cells[11]) == 4661                 # 전환단가
        assert round(cells[16], 4) == 0.0144            # 회원가입률
        assert round(cells[18], 4) == 0.051             # 결제완료율
        assert round(cells[20], 2) == 27.53             # 구매수익률(ROAS)

    def test_conversions_excluding_apply(self):
        cells = _prev_month_cells(PREV_TOTALS)
        # 신청제외 전환 = 총전환 7915 - 설명회신청 1119
        assert cells[12] == 6796

    def test_markup_cost_is_blank(self):
        # DB에 markup 광고비가 없다 — 옛 숫자를 남기느니 비운다
        assert _prev_month_cells(PREV_TOTALS)[8] is None

    def test_zero_denominators_do_not_raise(self):
        empty = dict.fromkeys(PREV_TOTALS, 0)
        cells = _prev_month_cells(empty)
        assert all(v == 0 for k, v in cells.items() if k != 8)


@needs_template
class TestSummaryPrevMonthRow:
    @staticmethod
    def _summary(prev_totals):
        return _export("26년 8월", 2026, 8, 31, prev_totals=prev_totals)["summary_26년 8월"]

    def test_prev_totals_are_written(self):
        ws = self._summary(PREV_TOTALS)

        assert ws.cell(PREV_MONTH_ROW, 2).value.strftime("%Y-%m") == "2026-07"
        assert ws.cell(PREV_MONTH_ROW, 3).value == 4574572
        assert ws.cell(PREV_MONTH_ROW, 7).value == 36887933.0
        assert ws.cell(PREV_MONTH_ROW, 8).value is None   # markup 광고비

    def test_without_prev_totals_row_is_blank(self):
        ws = self._summary(None)

        assert ws.cell(PREV_MONTH_ROW, 2).value.strftime("%Y-%m") == "2026-07"
        assert all(ws.cell(PREV_MONTH_ROW, c).value is None for c in range(3, 23))


# ── 시트 구성 ─────────────────────────────────────────────────────────────────


@needs_template
class TestSheetNaming:
    """템플릿의 PERIOD 토큰이 요청 기간으로 바뀌고, 그 참조까지 함께 따라와야 한다."""

    def test_only_requested_period_sheets_exist(self):
        wb = _export("26년 8월", 2026, 8, 31)

        assert all(name.endswith("26년 8월") for name in wb.sheetnames), wb.sheetnames
        assert len(wb.sheetnames) == len(_ALL_MEDIA) + 1  # 매체 5개 + summary

    def test_summary_is_the_sheet_excel_opens(self):
        assert _export("26년 8월", 2026, 8, 31).active.title == "summary_26년 8월"

    def test_no_placeholder_token_survives(self):
        """토큰이 남으면 시트 이름과 수식이 어긋나 Excel에서 #REF! 가 된다."""
        wb = _export("26년 8월", 2026, 8, 31)
        leftover = [
            (name, cell.coordinate)
            for name in wb.sheetnames
            for row in wb[name].iter_rows()
            for cell in row
            if isinstance(cell.value, str) and PLACEHOLDER in cell.value
        ]
        assert leftover == []

    def test_no_formula_references_a_missing_sheet(self):
        wb = _export("26년 8월", 2026, 8, 31)
        names = set(wb.sheetnames)

        dangling = [
            (sheet, cell.coordinate, ref)
            for sheet in wb.sheetnames
            for row in wb[sheet].iter_rows()
            for cell in row
            if isinstance(cell.value, str) and cell.value.startswith("=")
            for ref in re.findall(r"'([^']+)'!", cell.value)
            if ref not in names
        ]
        assert dangling == []

    def test_output_stays_small(self):
        kpis = {"네이버SA": _kpi_frame(2026, 8, 31)}
        raw = ExcelService().fill_template(kpis, "26년 8월", 2026, 8)
        # 예전 86MB 템플릿은 출력물이 87MB였다 — 슬림 템플릿 회귀 방지선
        assert len(raw) < 5 * 1024 * 1024


@needs_template
class TestExternalWorkbookLinks:
    """템플릿의 전년동월 행은 원래 외부 통합문서(=[1]네이버SA_7월!C22)를 참조했다.

    받는 사람에게 그 파일이 없어 빈칸이 되고, 이 행을 합산하는 summary 7행과 YOY 행까지
    함께 비었다. 이제 참조 자체가 템플릿에 없어야 한다.
    """

    def test_no_external_workbook_reference_survives(self):
        wb = _export("26년 8월", 2026, 8, 31)
        leftover = [
            (name, cell.coordinate)
            for name in wb.sheetnames
            for row in wb[name].iter_rows()
            for cell in row
            if isinstance(cell.value, str) and "[1]" in cell.value
        ]
        assert leftover == []

    def test_external_link_definition_is_dropped(self):
        """링크 정의만 남아도 Excel이 열 때마다 '연결된 데이터를 업데이트할까요?'를 묻는다."""
        kpis = {label: _kpi_frame(2026, 8, 31) for label in _ALL_MEDIA}
        raw = ExcelService().fill_template(kpis, "26년 8월", 2026, 8)

        with zipfile.ZipFile(io.BytesIO(raw)) as z:
            assert [n for n in z.namelist() if "externalLink" in n] == []
            assert "externalReferences" not in z.read("xl/workbook.xml").decode("utf-8")


# ── 매체 시트의 비교 행 (8=전년동월, 9=전월) ─────────────────────────────────


@needs_template
class TestMediaComparisonRows:
    def test_year_ago_values_are_written(self):
        ws = _export("26년 8월", 2026, 8, 31, yoy_media_totals={"네이버SA": _SAMPLE})["네이버SA_26년 8월"]

        assert ws.cell(YOY_ROW, 3).value == 111     # 노출
        assert ws.cell(YOY_ROW, 4).value == 22      # 클릭
        assert ws.cell(YOY_ROW, 7).value == 3300    # 광고비
        assert ws.cell(YOY_ROW, 14).value == 6      # 회원가입
        assert ws.cell(YOY_ROW, 16).value == 7      # 구매완료
        assert ws.cell(YOY_ROW, 18).value == 5      # 구매매출
        assert ws.cell(YOY_ROW, 21).value == 8      # 설명회신청

    def test_prev_month_values_are_written(self):
        """매체별 전월 행은 예전에 템플릿 상수 그대로여서 MOM 이 늘 같은 달과 비교됐다."""
        ws = _export("26년 8월", 2026, 8, 31, prev_media_totals={"네이버SA": _SAMPLE})["네이버SA_26년 8월"]

        assert ws.cell(PREV_ROW, 3).value == 111
        assert ws.cell(PREV_ROW, 7).value == 3300
        assert ws.cell(PREV_ROW, 18).value == 5

    def test_missing_media_is_filled_with_zero(self):
        # 1년 전 데이터가 아예 없는 매체 — 빈칸으로 두면 YOY가 조용히 비어버린다
        wb = _export("26년 8월", 2026, 8, 31, yoy_media_totals={"네이버SA": _SAMPLE})
        ws = wb["카카오SA_26년 8월"]

        assert [ws.cell(YOY_ROW, c).value for c in (3, 4, 7, 8, 11, 13, 15)] == [0] * 7

    def test_derived_cells_stay_as_formulas(self):
        """값 칸만 채우고 계산은 템플릿 수식에 맡긴다."""
        wb = _export("26년 8월", 2026, 8, 31, yoy_media_totals={"네이버SA": _SAMPLE})

        ws = wb["네이버SA_26년 8월"]
        assert ws.cell(YOY_ROW, 5).value == "=IFERROR(D8/C8,)"    # CTR
        assert ws.cell(YOY_ROW, 6).value == "=IFERROR(G8/D8,)"    # CPC
        assert ws.cell(YOY_ROW, 8).value == "=SUM(N8,P8,U8)"      # 총전환수
        assert ws.cell(PREV_ROW, 8).value == "=SUM(N9,P9,U9)"

        summary = wb["summary_26년 8월"]
        assert summary.cell(7, 3).value.startswith("=SUM(")        # 전년동월 합계
        assert summary.cell(10, 3).value == "=IFERROR(C9/C7-1,)"   # YOY

    def test_total_conversions_match_the_current_month_definition(self):
        """총전환수가 RAW 칸인 매체는 당월 수식과 같은 정의로 채운다.

        카카오는 당월이 회원가입+구매완료라, 전월/전년동월만 DB 총전환수를 쓰면
        같은 열에서 정의가 달라진 채로 MOM 이 계산된다.
        """
        wb = _export("26년 8월", 2026, 8, 31, yoy_media_totals={"카카오SA": _SAMPLE, "구글SA": _SAMPLE})

        assert wb["카카오SA_26년 8월"].cell(YOY_ROW, 8).value == 13   # 회원가입 6 + 구매완료 7
        assert wb["구글SA_26년 8월"].cell(YOY_ROW, 9).value == 21     # + 설명회신청 8


# ── 일별 구간 ─────────────────────────────────────────────────────────────────


@needs_template
class TestDailyRows:
    def test_raw_values_land_on_the_right_day(self):
        ws = _export("26년 8월", 2026, 8, 31)["네이버SA_26년 8월"]

        assert ws.cell(23, 2).value == datetime(2026, 8, 1)
        assert ws.cell(23, 3).value == 100     # 노출
        assert ws.cell(25, 3).value == 300     # 3일
        assert ws.cell(25, 7).value == 3000    # 광고비

    def test_days_without_data_stay_blank(self):
        """짧은 달의 남는 칸에 옛 숫자가 남으면 TOTAL이 조용히 부풀어 오른다."""
        ws = _export("26년 2월", 2026, 2, 28)["네이버SA_26년 2월"]

        for row in range(23 + 28, 54):
            assert [ws.cell(row, c).value for c in (3, 4, 7, 14, 16, 18, 21)] == [None] * 7

    def test_naver_bs_mobile_cost_column_is_empty(self):
        """브검MO(25열)에 상수가 남아 있어 매달 유령 광고비가 더해졌다 — 회귀 방지."""
        ws = _export("26년 8월", 2026, 8, 31)["네이버BS_26년 8월"]

        assert ws.cell(23, 24).value == 1000                        # 브검PC = DB 광고비
        assert all(ws.cell(r, 25).value is None for r in range(23, 54))

    def test_kakao_cost_has_no_hand_typed_leftovers(self):
        """카카오 광고비 열에는 '=70*1.1' 같은 손입력 수식이 남아 있었다."""
        ws = _export("26년 2월", 2026, 2, 28)["카카오SA_26년 2월"]

        assert all(ws.cell(r, 7).value is None for r in range(23 + 28, 54))


# ── summary 날짜·기준일 ───────────────────────────────────────────────────────


@needs_template
class TestSummaryDates:
    def test_sa_total_rows_show_year_ago_prev_and_current(self):
        ws = _export("26년 6월", 2026, 6, 30)["summary_26년 6월"]

        assert ws.cell(YOY_ROW - 1, 2).value == datetime(2025, 6, 1)     # 전년동월
        assert ws.cell(PREV_MONTH_ROW, 2).value == datetime(2026, 5, 1)  # 전월
        assert ws.cell(YOY_ROW + 1, 2).value == datetime(2026, 6, 1)     # 당월

    def test_january_rolls_back_to_previous_year(self):
        ws = _export("26년 1월", 2026, 1, 31)["summary_26년 1월"]

        assert ws.cell(YOY_ROW - 1, 2).value == datetime(2025, 1, 1)
        assert ws.cell(PREV_MONTH_ROW, 2).value == datetime(2025, 12, 1)
        assert ws.cell(YOY_ROW + 1, 2).value == datetime(2026, 1, 1)

    def test_daily_rows_cover_the_month(self):
        ws = _export("26년 6월", 2026, 6, 30)["summary_26년 6월"]

        assert ws.cell(SUMMARY_DAILY_ROW, 2).value == datetime(2026, 6, 1)
        assert ws.cell(SUMMARY_DAILY_ROW + 29, 2).value == datetime(2026, 6, 30)

    def test_unused_daily_slots_are_blank_in_short_months(self):
        """31칸 고정이라 30일 이하인 달은 남는 칸을 비워야 한다 — 안 비우면 옛 날짜가 남는다."""
        ws = _export("26년 2월", 2026, 2, 28)["summary_26년 2월"]

        leftover = [
            ws.cell(SUMMARY_DAILY_ROW + i, 2).value
            for i in range(28, SUMMARY_DAILY_SLOTS)
        ]
        assert leftover == [None, None, None]

    def test_month_length_and_end_date(self):
        ws = _export("26년 2월", 2026, 2, 28)["summary_26년 2월"]

        assert ws["D3"].value == 28
        assert ws["B1"].value == datetime(2026, 3, 1)   # 기간 종료일 = 다음 달 1일

    def test_finished_month_uses_month_end_as_base_date(self):
        """기준일이 =TODAY() 였을 때는 지난 달 리포트의 WoW 행이 통째로 비었다."""
        ws = _export("20년 3월", 2020, 3, 31)["summary_20년 3월"]

        assert ws["D1"].value == datetime(2020, 4, 1)

    def test_base_date_is_referenced_not_recomputed(self):
        """매체 시트의 전주/금주 날짜는 summary 기준일 한 칸만 보게 만들었다."""
        ws = _export("20년 3월", 2020, 3, 31)["네이버SA_20년 3월"]

        assert ws["B13"].value == "='summary_20년 3월'!$D$1-8"
        assert ws["B14"].value == "='summary_20년 3월'!$D$1-1"


# ── 매체별 예산 · 코멘트 ──────────────────────────────────────────────────────


@needs_template
class TestBudgetsAndComment:
    def test_budgets_are_written_per_media(self):
        ws = _export(
            "26년 8월", 2026, 8, 31,
            media_budgets={"네이버SA": 19_000_000, "카카오SA": 50_000},
        )["summary_26년 8월"]

        assert ws.cell(SUMMARY_BUDGET_ROWS["네이버SA"], SUMMARY_BUDGET_COL).value == 19_000_000
        assert ws.cell(SUMMARY_BUDGET_ROWS["카카오SA"], SUMMARY_BUDGET_COL).value == 50_000

    def test_unset_media_budget_is_zero(self):
        ws = _export("26년 8월", 2026, 8, 31, media_budgets={"네이버SA": 1})["summary_26년 8월"]

        assert ws.cell(SUMMARY_BUDGET_ROWS["구글SA"], SUMMARY_BUDGET_COL).value == 0

    def test_budget_subtotal_rows_stay_as_formulas(self):
        """SA total / SA+BS total / TOTAL 행은 건드리지 않는다."""
        ws = _export("26년 8월", 2026, 8, 31, media_budgets={"네이버SA": 1})["summary_26년 8월"]

        assert ws.cell(24, SUMMARY_BUDGET_COL).value == "=SUM(D21:D23)"
        assert ws.cell(26, SUMMARY_BUDGET_COL).value == "=SUM(D24,D25)"

    def test_comment_is_written(self):
        ws = _export("26년 8월", 2026, 8, 31, comment="CPC 상승")["summary_26년 8월"]
        assert ws["B32"].value == "CPC 상승"

    def test_without_comment_cell_is_blank(self):
        assert _export("26년 8월", 2026, 8, 31)["summary_26년 8월"]["B32"].value is None
